from __future__ import annotations

import io
from datetime import datetime, time

from django.contrib import admin, messages
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html

from .models import (
    Customer, Order, OrderItem, OrderNote, OrderStatusHistory,
    Promotion, Newsletter,
)
from .services.novapost import NovaPostService, NovaPostServiceError
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

# ─── Status colours ──────────────────────────────────────────────────────────

STATUS_COLORS = {
    'new':             ('#1565c0', '#e3f2fd'),
    'processing':      ('#e65100', '#fff3e0'),
    'pending_payment': ('#6a1b9a', '#f3e5f5'),
    'paid':            ('#2e7d32', '#e8f5e9'),
    'shipped':         ('#00838f', '#e0f7fa'),
    'delivered':       ('#558b2f', '#f1f8e9'),
    'completed':       ('#37474f', '#eceff1'),
    'cancelled':       ('#b71c1c', '#ffebee'),
    'refund':          ('#4e342e', '#efebe9'),
    'pending':         ('#455a64', '#eceff1'),
    'confirmed':       ('#1b5e20', '#e8f5e9'),
}


def colored_status(status: str, label: str) -> str:
    color, bg = STATUS_COLORS.get(status, ('#333', '#f5f5f5'))
    return format_html(
        '<span style="background:{};color:{};padding:3px 10px;border-radius:12px;'
        'font-size:12px;font-weight:600;white-space:nowrap">{}</span>',
        bg, color, label
    )


# ─── Inlines ─────────────────────────────────────────────────────────────────

class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    can_delete = False
    fields = ['get_product_link', 'product_name', 'product_sku', 'quantity', 'price', 'get_line_total']
    readonly_fields = ['get_product_link', 'product_name', 'product_sku', 'quantity', 'price', 'get_line_total']

    def has_add_permission(self, request, obj=None):
        return False

    def get_product_link(self, obj):
        if obj.product:
            url = f'/admin/products/product/{obj.product.pk}/change/'
            return format_html('<a href="{}">{}</a>', url, obj.get_display_name())
        return obj.get_display_name()
    get_product_link.short_description = 'Товар'

    def get_line_total(self, obj):
        return format_html('<strong>{} ₴</strong>', obj.get_cost())
    get_line_total.short_description = 'Сума'


class OrderStatusHistoryInline(admin.TabularInline):
    model = OrderStatusHistory
    extra = 0
    can_delete = False
    fields = ['created_at', 'old_status', 'new_status', 'changed_by', 'comment']
    readonly_fields = ['created_at', 'old_status', 'new_status', 'changed_by']
    ordering = ['-created_at']

    def has_add_permission(self, request, obj=None):
        return False


class OrderNoteInline(admin.StackedInline):
    model = OrderNote
    extra = 1
    fields = ['author', 'text', 'is_internal', 'created_at']
    readonly_fields = ['created_at']
    ordering = ['-created_at']


class CustomerOrderInline(admin.TabularInline):
    model = Order
    extra = 0
    can_delete = False
    fields = ['get_order_link', 'status', 'final_total', 'payment_method', 'is_paid', 'created_at']
    readonly_fields = ['get_order_link', 'status', 'final_total', 'payment_method', 'is_paid', 'created_at']

    def has_add_permission(self, request, obj=None):
        return False

    def get_order_link(self, obj):
        url = f'/admin/orders/order/{obj.pk}/change/'
        return format_html('<a href="{}">#{}</a>', url, obj.order_number)
    get_order_link.short_description = 'Замовлення'


# ─── Customer Admin ───────────────────────────────────────────────────────────

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = [
        'get_full_name_display', 'phone', 'email',
        'total_orders', 'get_total_spent_display',
        'tags', 'last_order_at',
    ]
    list_filter = ['last_order_at', 'first_order_at']
    search_fields = ['phone', 'first_name', 'last_name', 'email', 'tags']
    readonly_fields = ['total_orders', 'total_spent', 'first_order_at', 'last_order_at', 'created_at']
    ordering = ['-last_order_at']
    inlines = [CustomerOrderInline]
    actions = ['recalculate_stats']

    fieldsets = (
        ('Контактні дані', {
            'fields': ('phone', 'first_name', 'last_name', 'email')
        }),
        ('Статистика', {
            'fields': ('total_orders', 'total_spent', 'first_order_at', 'last_order_at'),
            'classes': ('collapse',),
        }),
        ('CRM', {
            'fields': ('tags', 'notes'),
        }),
    )

    def get_full_name_display(self, obj):
        name = obj.get_full_name()
        return format_html('<strong>{}</strong>', name)
    get_full_name_display.short_description = "Клієнт"

    def get_total_spent_display(self, obj):
        return format_html('<strong>{} ₴</strong>', obj.total_spent)
    get_total_spent_display.short_description = 'Загальна сума'

    def recalculate_stats(self, request, queryset):
        for customer in queryset:
            customer.recalculate_stats()
        self.message_user(request, f'Оновлено статистику для {queryset.count()} клієнтів')
    recalculate_stats.short_description = 'Оновити статистику'


# ─── Order Admin ──────────────────────────────────────────────────────────────

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = [
        'order_number', 'get_status_badge', 'get_customer_info',
        'get_total_display', 'get_payment_display',
        'delivery_method', 'created_at',
    ]
    list_filter = ['status', 'payment_method', 'delivery_method', 'is_paid', 'created_at']
    search_fields = ['order_number', 'first_name', 'last_name', 'phone', 'email', 'nova_poshta_ttn']
    date_hierarchy = 'created_at'
    ordering = ['-created_at']

    readonly_fields = [
        'order_number', 'created_at', 'updated_at',
        'first_name', 'last_name', 'patronymic', 'phone', 'email',
        'delivery_method', 'nova_poshta_city', 'nova_poshta_city_ref',
        'nova_poshta_warehouse', 'nova_poshta_warehouse_ref',
        'ukrposhta_city', 'ukrposhta_address', 'ukrposhta_index',
        'payment_method', 'payment_date', 'payment_intent_id',
        'subtotal_retail', 'product_discount', 'promo_code',
        'promo_discount', 'final_total', 'notes',
        'nova_poshta_ttn', 'get_ttn_link', 'customer',
    ]

    inlines = [OrderItemInline, OrderStatusHistoryInline, OrderNoteInline]
    actions = [
        'mark_processing', 'mark_paid', 'mark_shipped', 'mark_delivered',
        'mark_completed', 'mark_cancelled', 'mark_refund',
        'create_nova_poshta_ttn', 'export_excel',
    ]

    fieldsets = (
        ('Замовлення', {
            'fields': ('order_number', 'status', 'customer', 'created_at', 'updated_at')
        }),
        ('Клієнт', {
            'fields': ('first_name', 'last_name', 'patronymic', 'phone', 'email')
        }),
        ('Доставка', {
            'fields': (
                'delivery_method',
                ('nova_poshta_city', 'nova_poshta_city_ref'),
                ('nova_poshta_warehouse', 'nova_poshta_warehouse_ref'),
                'ukrposhta_city', 'ukrposhta_address', 'ukrposhta_index',
                ('nova_poshta_ttn', 'get_ttn_link'),
            )
        }),
        ('Ціни', {
            'fields': (
                'subtotal_retail', 'product_discount',
                'promo_code', 'promo_discount', 'final_total',
            )
        }),
        ('Оплата', {
            'fields': ('payment_method', 'is_paid', 'payment_date', 'payment_intent_id')
        }),
        ('Примітки', {
            'fields': ('notes', 'admin_notes'),
        }),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('customer').prefetch_related('items__product')

    # ── Display helpers ──────────────────────────────────────────────────────

    def get_status_badge(self, obj):
        label = obj.get_status_display()
        return colored_status(obj.status, label)
    get_status_badge.short_description = 'Статус'

    def get_customer_info(self, obj):
        name = obj.get_customer_name()
        phone = obj.phone
        return format_html('<strong>{}</strong><br><small>{}</small>', name, phone)
    get_customer_info.short_description = 'Клієнт'

    def get_total_display(self, obj):
        return format_html('<strong>{} ₴</strong>', obj.final_total)
    get_total_display.short_description = 'Сума'

    def get_payment_display(self, obj):
        label = 'Онлайн' if obj.payment_method == 'online' else 'Накладений'
        paid = '✓' if obj.is_paid else '✗'
        color = '#2e7d32' if obj.is_paid else '#b71c1c'
        return format_html(
            '{} <span style="color:{};font-weight:700">{}</span>',
            label, color, paid
        )
    get_payment_display.short_description = 'Оплата'

    def get_ttn_link(self, obj):
        if obj.nova_poshta_ttn:
            url = f"https://track.novaposhta.ua/uk?number={obj.nova_poshta_ttn}"
            return format_html('<a href="{}" target="_blank">{}</a>', url, obj.nova_poshta_ttn)
        return '-'
    get_ttn_link.short_description = 'Посилання на ТТН'

    # ── save_model: auto status history ─────────────────────────────────────

    def save_model(self, request, obj, form, change):
        old_status = ''
        if change and obj.pk:
            try:
                old_status = Order.objects.get(pk=obj.pk).status
            except Order.DoesNotExist:
                pass
        super().save_model(request, obj, form, change)
        if change and old_status and old_status != obj.status:
            OrderStatusHistory.objects.create(
                order=obj,
                old_status=old_status,
                new_status=obj.status,
                changed_by=str(request.user),
                comment='',
            )

    # ── Bulk actions ─────────────────────────────────────────────────────────

    def _bulk_status(self, request, queryset, status: str, label: str):
        for order in queryset:
            old = order.status
            if old != status:
                order.status = status
                order.save(update_fields=['status', 'updated_at'])
                OrderStatusHistory.objects.create(
                    order=order,
                    old_status=old,
                    new_status=status,
                    changed_by=str(request.user),
                )
        self.message_user(request, f'{label}: {queryset.count()} замовлень')

    def mark_processing(self, request, queryset):
        self._bulk_status(request, queryset, 'processing', 'В обробку')
    mark_processing.short_description = 'В обробку'

    def mark_paid(self, request, queryset):
        self._bulk_status(request, queryset, 'paid', 'Оплачено')
        queryset.update(is_paid=True)
    mark_paid.short_description = 'Оплачено'

    def mark_shipped(self, request, queryset):
        self._bulk_status(request, queryset, 'shipped', 'Відправлено')
    mark_shipped.short_description = 'Відправлено'

    def mark_delivered(self, request, queryset):
        self._bulk_status(request, queryset, 'delivered', 'Доставлено')
    mark_delivered.short_description = 'Доставлено'

    def mark_completed(self, request, queryset):
        self._bulk_status(request, queryset, 'completed', 'Завершено')
    mark_completed.short_description = 'Завершити'

    def mark_cancelled(self, request, queryset):
        self._bulk_status(request, queryset, 'cancelled', 'Скасовано')
    mark_cancelled.short_description = 'Скасувати'

    def mark_refund(self, request, queryset):
        self._bulk_status(request, queryset, 'refund', 'Повернення')
    mark_refund.short_description = 'Повернення'

    # ── Export Excel ─────────────────────────────────────────────────────────

    def export_excel(self, request, queryset):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self.message_user(request, 'Встановіть openpyxl', messages.ERROR)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Замовлення'

        headers = [
            '№ замовлення', 'Статус', 'Ім\'я', 'Прізвище', 'Телефон', 'Email',
            'Доставка', 'Місто', 'Відділення', 'ТТН',
            'Сума', 'Оплата', 'Оплачено', 'Дата',
        ]
        bold = Font(bold=True)
        fill = PatternFill('solid', fgColor='1a1a1a')
        wfont = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = wfont
            cell.fill = fill

        for row_idx, order in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=order.order_number)
            ws.cell(row=row_idx, column=2, value=order.get_status_display())
            ws.cell(row=row_idx, column=3, value=order.first_name)
            ws.cell(row=row_idx, column=4, value=order.last_name)
            ws.cell(row=row_idx, column=5, value=order.phone)
            ws.cell(row=row_idx, column=6, value=order.email)
            ws.cell(row=row_idx, column=7, value=order.get_delivery_method_display())
            ws.cell(row=row_idx, column=8, value=order.nova_poshta_city or order.ukrposhta_city)
            ws.cell(row=row_idx, column=9, value=order.nova_poshta_warehouse or order.ukrposhta_address)
            ws.cell(row=row_idx, column=10, value=order.nova_poshta_ttn)
            ws.cell(row=row_idx, column=11, value=float(order.final_total))
            ws.cell(row=row_idx, column=12, value=order.get_payment_method_display())
            ws.cell(row=row_idx, column=13, value='Так' if order.is_paid else 'Ні')
            ws.cell(row=row_idx, column=14, value=order.created_at.strftime('%d.%m.%Y %H:%M'))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
        return response
    export_excel.short_description = 'Експорт в Excel'

    # ── Nova Poshta TTN ───────────────────────────────────────────────────────

    def create_nova_poshta_ttn(self, request, queryset):
        if not getattr(settings, 'NOVAPOST_API_KEY', None):
            self.message_user(request, 'NOVAPOST_API_KEY не налаштований', messages.ERROR)
            return

        valid_orders = []
        for order in queryset:
            errs = []
            if order.delivery_method != 'nova_poshta':
                errs.append('Метод доставки не Nova Poshta')
            if not order.nova_poshta_city_ref:
                errs.append('Не вказаний REF міста')
            if not order.nova_poshta_warehouse_ref:
                errs.append('Не вказаний REF відділення')
            if order.nova_poshta_ttn:
                errs.append('ТТН вже створена')
            if errs:
                self.message_user(
                    request,
                    f'Замовлення #{order.order_number}: {"; ".join(errs)}',
                    messages.WARNING
                )
            else:
                valid_orders.append(order)

        if not valid_orders:
            return

        try:
            np_service = NovaPostService(settings.NOVAPOST_API_KEY)
            counterparty = np_service.get_counterparty()
            if not counterparty:
                self.message_user(request, 'Не знайдено контрагента відправника', messages.ERROR)
                return
            addresses = np_service.get_sender_addresses()
            contacts = np_service.get_sender_contacts()
            if not addresses or not contacts:
                self.message_user(request, 'Не можна отримати дані відправника', messages.ERROR)
                return

            sender_ref = counterparty.get('Ref')
            addr = addresses[0]
            sender_address_ref = addr.get('Ref')
            sender_city_ref = addr.get('CityRef')
            contact_ref = contacts[0].get('Ref')

            ok = fail = 0
            for order in valid_orders:
                try:
                    # Use weight from items if available
                    weight_g = 1000
                    if order.items.exists():
                        first_item = order.items.first()
                        if first_item and first_item.product and first_item.product.weight:
                            weight_g = int(sum(
                                float(item.product.weight or 1) * item.quantity * 1000
                                for item in order.items.all()
                                if item.product
                            ))
                            weight_g = max(weight_g, 100)

                    result = np_service.create_shipment(
                        recipient_city_ref=order.nova_poshta_city_ref,
                        recipient_warehouse_ref=order.nova_poshta_warehouse_ref,
                        recipient_name=order.get_customer_name(),
                        recipient_phone=order.phone,
                        sender_ref=sender_ref,
                        sender_city_ref=sender_city_ref,
                        sender_address_ref=sender_address_ref,
                        sender_contact_ref=contact_ref,
                        description=f'Замовлення #{order.order_number}',
                        cost=str(int(order.final_total)),
                        weight=str(weight_g),
                    )
                    if result.get('success') and result.get('data'):
                        ttn = result['data'][0].get('IntDocNumber')
                        if ttn:
                            order.nova_poshta_ttn = ttn
                            order.save(update_fields=['nova_poshta_ttn'])
                            ok += 1
                        else:
                            fail += 1
                    else:
                        errs = result.get('errors', ['Невідома помилка'])
                        self.message_user(request, f'#{order.order_number}: {"; ".join(errs)}', messages.WARNING)
                        fail += 1
                except NovaPostServiceError as e:
                    self.message_user(request, f'#{order.order_number}: {e}', messages.WARNING)
                    fail += 1

            if ok:
                self.message_user(request, f'Створено {ok} ТТН', messages.SUCCESS)
            if fail:
                self.message_user(request, f'Помилок: {fail}', messages.ERROR)

        except Exception as e:
            logger.exception(f'create_nova_poshta_ttn error: {e}')
            self.message_user(request, 'Внутрішня помилка', messages.ERROR)

    create_nova_poshta_ttn.short_description = 'Створити ТТН (Нова Пошта)'

    # ── Dashboard (changelist) ────────────────────────────────────────────────

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        response = super().changelist_view(request, extra_context=extra_context)
        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        today_utc = timezone.now().date()
        start_of_day = timezone.make_aware(
            datetime.combine(today_utc, time.min), timezone.utc
        )
        end_of_day = timezone.make_aware(
            datetime.combine(today_utc, time.max), timezone.utc
        )
        today_qs = Order.objects.filter(created_at__range=(start_of_day, end_of_day))

        status_counts = {}
        for status, label in Order.STATUS_CHOICES[:9]:  # only new statuses
            status_counts[status] = {
                'label': label,
                'count': qs.filter(status=status).count(),
                'color': STATUS_COLORS.get(status, ('#333', '#f5f5f5')),
            }

        response.context_data['crm_summary'] = {
            'status_counts': status_counts,
            'today_count': today_qs.count(),
            'today_sum': today_qs.exclude(status='cancelled').aggregate(
                s=Sum('final_total')
            )['s'] or 0,
            'unpaid': qs.filter(is_paid=False).exclude(status__in=['cancelled', 'refund']).count(),
            'no_ttn': qs.filter(
                delivery_method='nova_poshta',
                nova_poshta_ttn='',
                status__in=['paid', 'processing', 'shipped'],
            ).count(),
        }
        return response


# ─── Promotion Admin ──────────────────────────────────────────────────────────

@admin.register(Promotion)
class PromotionAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'get_discount_display', 'apply_to', 'get_usage', 'get_status', 'is_active', 'start_date', 'end_date']
    list_filter = ['is_active', 'discount_type', 'apply_to', 'start_date']
    search_fields = ['name', 'code']
    readonly_fields = ['uses_count', 'created_at']
    filter_horizontal = ['categories']
    date_hierarchy = 'start_date'
    actions = ['activate_promotions', 'deactivate_promotions', 'duplicate_promo']

    fieldsets = (
        ('Основна інформація', {'fields': ('name', 'code', 'is_active')}),
        ('Умови знижки', {'fields': (('discount_type', 'discount_value'), 'min_order_amount')}),
        ('Застосування', {'fields': ('apply_to', 'categories')}),
        ('Термін дії', {'fields': (('start_date', 'end_date'),)}),
        ('Обмеження', {'fields': (('max_uses', 'uses_count'),)}),
    )

    def get_discount_display(self, obj):
        if obj.discount_type == 'percentage':
            return format_html('<strong>{}%</strong>', obj.discount_value)
        return format_html('<strong>{} ₴</strong>', obj.discount_value)
    get_discount_display.short_description = 'Знижка'

    def get_usage(self, obj):
        if obj.max_uses:
            pct = (obj.uses_count / obj.max_uses) * 100
            color = '#4caf50' if pct < 80 else '#ff9800' if pct < 100 else '#f44336'
            return format_html('<span style="color:{};font-weight:600">{}/{}</span>', color, obj.uses_count, obj.max_uses)
        return format_html('<span style="color:#2196f3">{}</span>', obj.uses_count)
    get_usage.short_description = 'Використань'

    def get_status(self, obj):
        if obj.is_valid():
            return format_html('<span style="color:#4caf50;font-weight:600">✓ Активний</span>')
        elif not obj.is_active:
            return format_html('<span style="color:#999">✗ Вимкнено</span>')
        return format_html('<span style="color:#ff9800">⏰ Прострочений</span>')
    get_status.short_description = 'Статус'

    def activate_promotions(self, request, queryset):
        queryset.update(is_active=True)
    activate_promotions.short_description = 'Активувати'

    def deactivate_promotions(self, request, queryset):
        queryset.update(is_active=False)
    deactivate_promotions.short_description = 'Деактивувати'

    def duplicate_promo(self, request, queryset):
        for p in queryset:
            p.pk = None
            p.code = f'{p.code}_copy'
            p.uses_count = 0
            p.save()
    duplicate_promo.short_description = 'Дублювати'


# ─── Newsletter Admin ─────────────────────────────────────────────────────────

@admin.register(Newsletter)
class NewsletterAdmin(admin.ModelAdmin):
    list_display = ['email', 'is_active', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['email']
    list_editable = ['is_active']
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'
    actions = ['export_emails']

    def export_emails(self, request, queryset):
        emails = list(queryset.filter(is_active=True).values_list('email', flat=True))
        content = '\n'.join(emails)
        response = HttpResponse(content, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="subscribers.txt"'
        return response
    export_emails.short_description = 'Експортувати email'
